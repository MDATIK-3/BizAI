import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_headers(ws, row_num):
    """Apply header styling to a row"""
    for cell in ws[row_num]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

# ============ CUSTOMERS SHEET ============
ws_customers = wb.create_sheet("Customers")
customer_headers = ["Customer ID", "Name", "Email", "Phone", "Category", "Total Purchases", "Last Purchase Date", "Purchase Frequency", "Status"]
ws_customers.append(customer_headers)
style_headers(ws_customers, 1)

customer_data = [
    (101, "Dhaka Fashion Store", "info@dhakafashion.com", "01700123456", "Retail", 250000, "2025-11-01", "Weekly", "Active"),
    (102, "Sylhet Textiles", "contact@sylhettex.com", "01800234567", "Wholesale", 450000, "2025-10-28", "Bi-weekly", "Active"),
    (103, "Chattogram Electronics", "hello@chattelec.com", "01900345678", "Retail", 125000, "2025-09-15", "Monthly", "At-Risk"),
    (104, "Rajshahi Food Supplies", "orders@rajfood.com", "02000456789", "Wholesale", 600000, "2025-11-02", "Weekly", "Active"),
    (105, "Narayanganj Gifts", "sales@nrgifts.com", "02100567890", "Retail", 80000, "2025-08-20", "Monthly", "Inactive"),
    (106, "Barisal Hardware", "barisal.hw@gmail.com", "02200678901", "Retail", 95000, "2025-10-10", "Monthly", "At-Risk"),
    (107, "Khulna Garments", "khulna.gmt@gmail.com", "02300789012", "Wholesale", 520000, "2025-11-03", "Weekly", "Active"),
    (108, "Mymensingh Traders", "traders@mymensingh.com", "02400890123", "Wholesale", 380000, "2025-10-30", "Bi-weekly", "Active"),
]

for customer in customer_data:
    ws_customers.append(customer)

ws_customers.column_dimensions['A'].width = 12
ws_customers.column_dimensions['B'].width = 22
ws_customers.column_dimensions['C'].width = 25
ws_customers.column_dimensions['D'].width = 13
ws_customers.column_dimensions['E'].width = 12
ws_customers.column_dimensions['F'].width = 15
ws_customers.column_dimensions['G'].width = 18
ws_customers.column_dimensions['H'].width = 16
ws_customers.column_dimensions['I'].width = 12

# ============ SALES TRANSACTIONS SHEET ============
ws_sales = wb.create_sheet("Sales Transactions")
sales_headers = ["Transaction ID", "Date", "Customer ID", "Product", "Quantity", "Unit Price", "Total", "Payment Method", "Status"]
ws_sales.append(sales_headers)
style_headers(ws_sales, 1)

products = ["T-Shirts", "Jeans", "Sarees", "Shirts", "Trousers", "Electronics", "Shoes", "Accessories"]
payment_methods = ["Cash", "Bank Transfer", "Mobile Banking", "Check"]
today = datetime.now()

for i in range(50):
    trans_id = 5001 + i
    date = (today - timedelta(days=random.randint(0, 60))).strftime("%Y-%m-%d")
    customer_id = random.choice([101, 102, 103, 104, 105, 106, 107, 108])
    product = random.choice(products)
    quantity = random.randint(5, 100)
    unit_price = random.randint(200, 5000)
    total = quantity * unit_price
    payment = random.choice(payment_methods)
    status = random.choice(["Completed", "Completed", "Pending", "Completed"])
    
    ws_sales.append([trans_id, date, customer_id, product, quantity, unit_price, total, payment, status])

ws_sales.column_dimensions['A'].width = 14
ws_sales.column_dimensions['B'].width = 12
ws_sales.column_dimensions['C'].width = 12
ws_sales.column_dimensions['D'].width = 16
ws_sales.column_dimensions['E'].width = 10
ws_sales.column_dimensions['F'].width = 11
ws_sales.column_dimensions['G'].width = 10
ws_sales.column_dimensions['H'].width = 16
ws_sales.column_dimensions['I'].width = 10

# ============ INVENTORY SHEET ============
ws_inventory = wb.create_sheet("Inventory")
inventory_headers = ["SKU", "Product Name", "Category", "Current Stock", "Reorder Level", "Last Updated", "Location", "Cost Price", "Selling Price"]
ws_inventory.append(inventory_headers)
style_headers(ws_inventory, 1)

inventory_data = [
    ("SKU001", "T-Shirts", "Clothing", 450, 100, "2025-11-03", "Warehouse A", 150, 400),
    ("SKU002", "Jeans", "Clothing", 280, 80, "2025-11-02", "Warehouse A", 300, 800),
    ("SKU003", "Sarees", "Clothing", 120, 50, "2025-11-01", "Warehouse B", 500, 1500),
    ("SKU004", "Shirts", "Clothing", 350, 100, "2025-11-03", "Warehouse A", 200, 600),
    ("SKU005", "Trousers", "Clothing", 200, 60, "2025-10-31", "Warehouse B", 250, 700),
    ("SKU006", "Shoes", "Footwear", 95, 40, "2025-11-01", "Warehouse C", 400, 1200),
    ("SKU007", "Accessories", "Accessories", 500, 150, "2025-11-03", "Warehouse A", 50, 200),
    ("SKU008", "Electronics", "Electronics", 45, 20, "2025-10-30", "Warehouse C", 2000, 5000),
]

for item in inventory_data:
    ws_inventory.append(item)

ws_inventory.column_dimensions['A'].width = 10
ws_inventory.column_dimensions['B'].width = 16
ws_inventory.column_dimensions['C'].width = 12
ws_inventory.column_dimensions['D'].width = 13
ws_inventory.column_dimensions['E'].width = 13
ws_inventory.column_dimensions['F'].width = 14
ws_inventory.column_dimensions['G'].width = 12
ws_inventory.column_dimensions['H'].width = 11
ws_inventory.column_dimensions['I'].width = 13

# ============ CASH FLOW SHEET ============
ws_cashflow = wb.create_sheet("Cash Flow")
cashflow_headers = ["Date", "Description", "Type", "Amount", "Balance", "Category", "Reference"]
ws_cashflow.append(cashflow_headers)
style_headers(ws_cashflow, 1)

balance = 500000
for i in range(30):
    date = (today - timedelta(days=30-i)).strftime("%Y-%m-%d")
    
    if i % 3 == 0:
        amount = random.randint(5000, 25000)
        type_val = "Inflow"
        description = f"Sales Revenue - {random.choice(products)}"
        category = "Revenue"
        reference = f"INV-{5001+i}"
    else:
        amount = random.randint(2000, 15000)
        type_val = "Outflow"
        description = random.choice(["Material Purchase", "Salary Payment", "Utilities", "Rent", "Logistics"])
        category = random.choice(["Expenses", "Payroll", "Supplies"])
        reference = f"EXP-{3001+i}"
    
    if type_val == "Inflow":
        balance += amount
    else:
        balance -= amount
    
    ws_cashflow.append([date, description, type_val, amount, balance, category, reference])

ws_cashflow.column_dimensions['A'].width = 12
ws_cashflow.column_dimensions['B'].width = 22
ws_cashflow.column_dimensions['C'].width = 10
ws_cashflow.column_dimensions['D'].width = 12
ws_cashflow.column_dimensions['E'].width = 12
ws_cashflow.column_dimensions['F'].width = 14
ws_cashflow.column_dimensions['G'].width = 14

# ============ PRODUCT DEMAND SHEET ============
ws_demand = wb.create_sheet("Product Demand")
demand_headers = ["Date", "Product", "Units Sold", "Peak Hours", "Avg Rating", "Returns", "Season"]
ws_demand.append(demand_headers)
style_headers(ws_demand, 1)

seasons = ["Summer", "Monsoon", "Winter"]
for i in range(60):
    date = (today - timedelta(days=60-i)).strftime("%Y-%m-%d")
    product = random.choice(products)
    units = random.randint(10, 150)
    peak_hours = random.choice(["9AM-12PM", "2PM-5PM", "6PM-9PM", "All Day"])
    rating = round(random.uniform(3.5, 5.0), 1)
    returns = random.randint(0, units // 20)
    season = random.choice(seasons)
    
    ws_demand.append([date, product, units, peak_hours, rating, returns, season])

ws_demand.column_dimensions['A'].width = 12
ws_demand.column_dimensions['B'].width = 14
ws_demand.column_dimensions['C'].width = 12
ws_demand.column_dimensions['D'].width = 14
ws_demand.column_dimensions['E'].width = 11
ws_demand.column_dimensions['F'].width = 9
ws_demand.column_dimensions['G'].width = 10

# Save file
wb.save('demo_business_data.xlsx')
print("[v0] Demo Excel file created successfully: demo_business_data.xlsx")
print("[v0] Sheets created: Customers, Sales Transactions, Inventory, Cash Flow, Product Demand")
